import openpyxl, os, pprint, re

##########################
#Sätt upp REGEX			 #
##########################
reg_swe = "SE$"
reg_nor = "NO$"
reg_den = "DK$"
reg_fin = "FI$"
reg_rap = "Rapport"
reg_file = "^20[\d]+ sergelOutput.xlsx$"

# Lista de filer som finns i katalogen
dir = os.listdir()
print('Filer i katalogen:')
print()
print(dir)
print()
# Gå igenom varje fil i katalogen tills rätt fil hittas
# Rätt fil är 20YYMM sergelOutput.xlsx
found = False
for file in dir:
	if re.search(reg_file, file):
		file_name = file
		print("Hittade filen " + file)
		found = True
		break
# Hittade vi inte en fil som matchar namnet vi letar efter avsluta programmet.
if found is False:
	print('Hittade ingen fil med namnet "20YYMM sergelOutput.xlsx"')
	print('Kontrollera om du står i rätt katalog eller om filen har bytt namn')
	input('Tryck på en knapp för att avsluta programmet')
	exit(0)


# Öppna filen som data ska läsas från
wb = openpyxl.load_workbook(file_name)
sheet = wb['sergelOutput']


print('Öppnat excel')

# Skapa den "dictionary" som data ska lagras i tillfälligt.
dataSet = {}

print('Läser in excel')
##############################################
# Iterera över alla rader i den inlästa filen.
# Skriv in data i dataSet-dictionaryn.
# setdefault lägger till data till en matchande befintlig entry.
# Finns det ingen matchande entry skapas en ny.
# På detta sätt skapas alla kunder, användare 
# och SMS dynamiskt.
# Det här behöver alltså aldrig ändras om inte output-filen ändras
##############################################
for row in range(1, sheet.max_row + 1):
	customer = sheet['A' + str(row)].value
	dataSet.setdefault(customer, {})
	
	user = sheet['B' + str(row)].value
	dataSet[customer].setdefault(user, {})
	
	service = sheet['C' + str(row)].value
	dataSet[customer][user].setdefault(service, {'NUM': 0})
	
	dataSet[customer][user][service]['NUM'] = sheet['D' + str(row)].value

print('Läst in data')

##########################
#Skriv data till ny excel#
##########################

#Skriv översta raden som ska användas för filter
sheet.cell(row = 1, column = 1).value = 'KUND'
sheet.cell(row = 1, column = 2).value = 'ANVÄNDARE'
sheet.cell(row = 1, column = 3).value = 'TJÄNST'
sheet.cell(row = 1, column = 4).value = 'ANTAL'
sheet.cell(row = 1, column = 5).value = 'SWE'
sheet.cell(row = 1, column = 6).value = 'NOR'
sheet.cell(row = 1, column = 7).value = 'DEN'
sheet.cell(row = 1, column = 8).value = 'FIN'
sheet.cell(row = 1, column = 9).value = 'OTHER'

print('Skriver resultat')

# Iterera över all data som finns och skriv den till excel
# "i" är en iterator som håller reda på vilken rad det ska skrivas till
# Börjar på två för att inte skriva över filter-raden
#
# OBS openpyxl är inte nollindexerat OBS
#
i = 2
for customer in dataSet:
	for user in dataSet[customer]:
		for service in dataSet[customer][user]:
			sheet.cell(row = i, column = 1).value = customer
			sheet.cell(row = i, column = 2).value = user
			sheet.cell(row = i, column = 3).value = service
			sheet.cell(row = i, column = 4).value = dataSet[customer][user][service]['NUM']
			
			# Matcha "service" med rätt land och skriv antal i motsvarande kolumn
			if re.search(reg_swe,service):
				sheet.cell(row = i, column = 5).value = dataSet[customer][user][service]['NUM']
			elif re.search(reg_nor,service):
				sheet.cell(row = i, column = 6).value = dataSet[customer][user][service]['NUM']
			elif re.search(reg_den,service):
				sheet.cell(row = i, column = 7).value = dataSet[customer][user][service]['NUM']
			elif re.search(reg_fin,service):
				sheet.cell(row = i, column = 8).value = dataSet[customer][user][service]['NUM']
			#Rapport visas inte som 
			elif re.search(reg_rap,service):
				pass #DO NOTHING
			else:
				sheet.cell(row = i, column = 9).value = dataSet[customer][user][service]['NUM']
			
			i += 1

# Spara den modifierade filen med "_PY" på slutet
new_file_name = file_name.replace(".xlsx", "_PY.xlsx")
wb.save(new_file_name)
print('Klar!')
input('Tryck på en knapp för att avsluta programmet')

